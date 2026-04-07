from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import numbers

LOCAL_TZ = timezone.get_default_timezone()
TIME_FORMAT = '[hh]:mm'

def get_full_name(user):
    full_name = ''
    if user.first_name:
        full_name += user.first_name + ' '
    if user.last_name:
        full_name += user.last_name
    if full_name == '':
        full_name += user.username
    return full_name


def excel_export(modeladmin, request, queryset):
    # Crea un libro de Excel y selecciona la hoja activa
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Entradas"
    bold_font = Font(bold=True)

    # Define los encabezados
    headers = ['Usuario', 'Fecha', 'Hora Inicio', 'Hora Fin', 'Duración']
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.font = bold_font
    worksheet.column_dimensions['A'].width = 16
    worksheet.column_dimensions['B'].width = 16


    current_user = None
    user_start_row = None  # primera fila del usuario actual
    totals = []

    for workperiod in queryset:
        # ---- Hoja DETALLE ----
        start = workperiod.start_time
        end = workperiod.end_time
        user = workperiod.user

        if not all([start, end, user]):
            continue

        # ---- Detectar cambio de usuario ----
        if current_user is not None and user.id != current_user.id:
            last_row = worksheet.max_row
            print(last_row)

            # Fila TOTAL del usuario anterior
            total_row = last_row + 1
            worksheet.cell(row=total_row, column=1,
                           value=f"Total {get_full_name(current_user)}")
            total_cell = worksheet.cell(row=total_row, column=5)
            total_cell.value = f"=SUM(E{user_start_row}:E{last_row})"
            total_cell.number_format = TIME_FORMAT
            totals.append((
                get_full_name(current_user),
                total_row
            ))

            for cell in worksheet[total_row]:
                cell.font = bold_font

            # Fila en blanco entre usuarios
            worksheet.append([])

            user_start_row = None

        current_user = user

        # ---- Manejo de timezones ----
        if timezone.is_naive(start):
            start = timezone.make_aware(start, timezone=timezone.utc)
        if timezone.is_naive(end):
            end = timezone.make_aware(end, timezone=timezone.utc)

        # Convertir a la TZ del proyecto (aware)
        start = timezone.localtime(start, LOCAL_TZ) if start else None
        end = timezone.localtime(end, LOCAL_TZ) if end else None

        # Excel NO soporta tzinfo: convertir a naive manteniendo la hora local
        # Y ajustar a 0 segundos
        if start:
            start = start.replace(tzinfo=None, second=0)
        if end:
            end = end.replace(tzinfo=None, second=0)

        worksheet.append([
            get_full_name(user),
            start.date(),
            start,
            end,
            None
        ])

        current_row = worksheet.max_row
        if user_start_row is None:
            user_start_row = current_row

        # Formatos de la fila recién añadida
        worksheet.cell(row=current_row, column=2).number_format = numbers.FORMAT_DATE_YYYYMMDD2
        worksheet.cell(row=current_row, column=3).number_format = 'hh:mm'
        worksheet.cell(row=current_row, column=4).number_format = 'hh:mm'

        duration_cell = worksheet.cell(row=current_row, column=5)
        duration_cell.value = f"=D{current_row}-C{current_row}"
        duration_cell.number_format = TIME_FORMAT

    # ---- TOTAL del último usuario ----
    if current_user and user_start_row:
        last_row = worksheet.max_row
        total_row = last_row + 1

        worksheet.cell(row=total_row, column=1,
                       value=f"Total {get_full_name(current_user)}")
        total_cell = worksheet.cell(row=total_row, column=5)
        total_cell.value = f"=SUM(E{user_start_row}:E{last_row})"
        total_cell.number_format = TIME_FORMAT

        totals.append((
            get_full_name(current_user),
            total_row
        ))

        for cell in worksheet[total_row]:
            cell.font = bold_font

    # ---- Hoja RESUMEN ----
    summary = workbook.create_sheet(title="Resumen")

    summary.append(['Usuario', 'Total'])
    for cell in summary[1]:
        cell.font = bold_font

    summary.column_dimensions['A'].width = 20
    summary.column_dimensions['B'].width = 15

    for row_idx, (user_name, total_row) in enumerate(totals, start=2):
        summary.cell(row=row_idx, column=1, value=user_name)

        total_cell = summary.cell(row=row_idx, column=2)
        total_cell.value = f"=Entradas!E{total_row}"
        total_cell.number_format = TIME_FORMAT

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=entradas.xlsx'
    workbook.save(response)
    return response

excel_export.short_description = "Exportar entradas a Excel"
